import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import MergedCell
import math
import os
import json

# --- JSON data file ---
DATA_FILE = "break_schedule_data.json"

# --- Page setup ---
st.set_page_config(page_title="Break Scheduler", layout="wide")
st.title("☕ Break Scheduler with Checker")

# --- Settings ---
break15 = timedelta(minutes=15)
break30 = timedelta(minutes=30)
stagger_gap = timedelta(minutes=0)

# --- Load existing data ---
if os.path.exists(DATA_FILE):
    with open(DATA_FILE, "r") as f:
        saved_data = json.load(f)
    
    st.session_state['tables'] = {giver: pd.DataFrame(df) for giver, df in saved_data.get("tables", {}).items()}
    form_data = saved_data.get("form_data", {})
    givers_input_default = form_data.get("givers_input", "1,2,3,4")
    shift_A_input_default = form_data.get("shift_A_input", "1,2,3,4,5,6,7,8")
    shift_B_input_default = form_data.get("shift_B_input", "1b,2b,3b,4b,5b,6b,7b,8b")
    giver_max_breaks_default = form_data.get("giver_max_breaks", {})
    giver_break_type_default = form_data.get("giver_break_type", {})
    giver_shift_times_default = {}
    for g, times in form_data.get("giver_shift_times", {}).items():
        start = datetime.strptime(times[0], "%H:%M:%S").time()
        end = datetime.strptime(times[1], "%H:%M:%S").time()
        giver_shift_times_default[g] = (start, end)
    B_shift_start_time_default = datetime.strptime(form_data.get("B_shift_start_time", "13:00:00"), "%H:%M:%S").time()
    schedule_date_default = datetime.strptime(form_data.get("schedule_date", str(datetime.today().date())), "%Y-%m-%d").date()
else:
    st.session_state['tables'] = {}
    givers_input_default = "1,2,3,4"
    shift_A_input_default = "1,2,3,4,5,6,7,8"
    shift_B_input_default = "1b,2b,3b,4b,5b,6b,7b,8b"
    giver_max_breaks_default = {}
    giver_break_type_default = {}
    giver_shift_times_default = {}
    B_shift_start_time_default = datetime.strptime("13:00", "%H:%M").time()
    schedule_date_default = datetime.today().date()

# --- Inputs ---
st.subheader("👨‍💼 Break Giver(s)")
givers_input = st.text_input("Enter break giver names (comma-separated)", givers_input_default)
givers = [g.strip() for g in givers_input.split(",") if g.strip()]

st.subheader("👥 Employees")
shift_A_input = st.text_area("Shift A Employees (comma-separated)", shift_A_input_default)
shift_B_input = st.text_area("Shift B Employees (comma-separated)", shift_B_input_default)
shift_employees = {
    "A": [e.strip() for e in shift_A_input.split(",") if e.strip()],
    "B": [e.strip() for e in shift_B_input.split(",") if e.strip()]
}

# --- Breaker max breaks and type ---
st.subheader("Assign number of breaks per Breaker and Break Type")
giver_max_breaks = {}
giver_break_type = {}
cols = st.columns(len(givers))
for i, giver in enumerate(givers):
    with cols[i]:
        # Number of breaks
        default_val = giver_max_breaks_default.get(giver, 4)
        giver_max_breaks[giver] = st.number_input(f"Breaks for {giver}", min_value=1, max_value=20, value=default_val, step=1)
        
        # Break type
        giver_break_type[giver] = st.selectbox(
            f"Break Type for {giver}",
            options=["15 min only", "30 min only", "Both"],
            index=["15 min only", "30 min only", "Both"].index(
                giver_break_type_default.get(giver, "Both")
            )
        )

# --- Shift input per giver ---
st.subheader("Breaker Shift Times")
giver_shift_times = {}
for giver in givers:
    col1, col2 = st.columns(2)
    with col1:
        default_start = giver_shift_times_default.get(giver, (datetime.strptime("09:00","%H:%M").time(), None))[0]
        start_time = st.time_input(f"{giver} Shift Start", default_start)
    with col2:
        default_end = giver_shift_times_default.get(giver, (None, datetime.strptime("17:00","%H:%M").time()))[1]
        end_time = st.time_input(f"{giver} Shift End", default_end)
    giver_shift_times[giver] = (start_time, end_time)

# --- B-Shift start time ---
st.subheader("B-Shift Timing")
B_shift_start_time = st.time_input("B Shift Start Time (breaks start 1 hour after)", B_shift_start_time_default)

# --- Schedule date ---
schedule_date = st.date_input("📅 Select Schedule Date", schedule_date_default)

# --- Generate Schedule ---
generate = st.button("Generate Schedule")

if generate:
    st.session_state['tables'] = {}
    A_queue = shift_employees["A"].copy()
    B_queue_all = shift_employees["B"].copy()

    for giver in givers:
        max_breaks = giver_max_breaks[giver]
        break_type = giver_break_type[giver]

        # Assign employees
        assigned_A = A_queue.copy()
        assigned_B = B_queue_all.copy()
        schedule = []
        current_time = datetime.combine(schedule_date, giver_shift_times[giver][0])
        breaks_assigned = 0

        # Helper function for round-robin
        def get_next_employee(queue):
            if not queue:
                return None
            emp = queue.pop(0)
            queue.append(emp)
            return emp

        while breaks_assigned < max_breaks:
            # A-shift 15 min
            if break_type in ["15 min only", "Both"] and assigned_A:
                emp = get_next_employee(assigned_A)
                start = current_time
                end = start + break15
                schedule.append([emp, "15 min", start.strftime("%H:%M"), end.strftime("%H:%M"), ""])
                current_time = end + stagger_gap
                breaks_assigned += 1
                if breaks_assigned >= max_breaks:
                    break

            # Giver self-break (30 min)
            if break_type in ["30 min only", "Both"] and breaks_assigned == 3 and max_breaks >= 4:
                start = current_time
                end = start + break30
                schedule.append([giver, "30 min (Giver)", start.strftime("%H:%M"), end.strftime("%H:%M"), ""])
                current_time = end + stagger_gap
                breaks_assigned += 1
                if breaks_assigned >= max_breaks:
                    break

            # A-shift 30 min
            if break_type in ["30 min only", "Both"] and assigned_A:
                emp = get_next_employee(assigned_A)
                start = current_time
                end = start + break30
                schedule.append([emp, "30 min", start.strftime("%H:%M"), end.strftime("%H:%M"), ""])
                current_time = end + stagger_gap
                breaks_assigned += 1
                if breaks_assigned >= max_breaks:
                    break

            # B-shift
            if assigned_B:
                B_shift_min_start = datetime.combine(schedule_date, B_shift_start_time) + timedelta(hours=1)
                if current_time < B_shift_min_start:
                    current_time = B_shift_min_start

                if break_type in ["15 min only", "Both"]:
                    emp = get_next_employee(assigned_B)
                    start = current_time
                    end = start + break15
                    schedule.append([emp, "15 min", start.strftime("%H:%M"), end.strftime("%H:%M"), ""])
                    current_time = end + stagger_gap
                    breaks_assigned += 1
                    if breaks_assigned >= max_breaks:
                        break

                if break_type in ["30 min only", "Both"]:
                    emp = get_next_employee(assigned_B)
                    start = current_time
                    end = start + break30
                    schedule.append([emp, "30 min", start.strftime("%H:%M"), end.strftime("%H:%M"), ""])
                    current_time = end + stagger_gap
                    breaks_assigned += 1
                    if breaks_assigned >= max_breaks:
                        break

        # Total time
        if schedule:
            first_start = datetime.strptime(schedule[0][2], "%H:%M")
            last_end = datetime.strptime(schedule[-1][3], "%H:%M")
            total_time = last_end - first_start
            schedule.append(["", "Total Time", first_start.strftime("%H:%M"), last_end.strftime("%H:%M"), str(total_time)])

        df = pd.DataFrame(schedule, columns=["Employee", "Break Type", "Start", "End", "SA Initial"])
        st.session_state['tables'][giver] = df

    # --- Editable Tables ---
    st.subheader("📅 Editable Schedule Per Break Giver")
    for giver, df in st.session_state['tables'].items():
        start_time, end_time = giver_shift_times.get(
            giver, 
            (datetime.strptime("09:00","%H:%M").time(), datetime.strptime("17:00","%H:%M").time())
        )
        st.markdown(f"**Breaker: {giver} | Date: {schedule_date} | Start: {start_time} | End: {end_time}**")
        edited_df = st.data_editor(df, num_rows="dynamic", use_container_width=True, key=f"editor_{giver}")
        st.session_state['tables'][giver] = edited_df

    # --- Save all data to JSON ---
    to_save = {
        "tables": {giver: df.to_dict(orient="records") for giver, df in st.session_state['tables'].items()},
        "form_data": {
            "givers_input": givers_input,
            "shift_A_input": shift_A_input,
            "shift_B_input": shift_B_input,
            "giver_max_breaks": giver_max_breaks,
            "giver_break_type": giver_break_type,
            "giver_shift_times": {g: [str(t[0]), str(t[1])] for g, t in giver_shift_times.items()},
            "B_shift_start_time": str(B_shift_start_time),
            "schedule_date": str(schedule_date)
        }
    }
    with open(DATA_FILE, "w") as f:
        json.dump(to_save, f, indent=2)

    st.success("✅ Schedule generated, displayed, and saved successfully!")

# --- Break Counter ---
st.subheader("📝 Break Count Per Employee")
if 'tables' in st.session_state:
    all_rows = pd.concat(st.session_state['tables'].values(), ignore_index=True)
    
    counter_list = []
    for emp in shift_employees["A"] + shift_employees["B"]:
        count_A = len(all_rows[(all_rows["Employee"] == emp) & (all_rows["Employee"].isin(shift_employees["A"]))])
        count_B = len(all_rows[(all_rows["Employee"] == emp) & (all_rows["Employee"].isin(shift_employees["B"]))])
        counter_list.append([emp, count_A, count_B])
    
    counter_df = pd.DataFrame(counter_list, columns=["Employee", "Shift A Breaks", "Shift B Breaks"])
    st.dataframe(counter_df)

# --- Excel Export (Single Sheet) ---
st.subheader("⬇️ Download Schedule (Single Sheet)")

buffer = BytesIO()
wb = Workbook()
ws = wb.active
ws.title = "Schedule"

if 'tables' in st.session_state:
    for giver, df in st.session_state['tables'].items():
        start_time, end_time = giver_shift_times.get(
            giver, 
            (datetime.strptime("09:00","%H:%M").time(), datetime.strptime("17:00","%H:%M").time())
        )
        # Table title
        ws.append([f"Breaker: {giver} | Date: {schedule_date} | Start: {start_time} | End: {end_time}"])
        title_row = ws.max_row
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=df.shape[1])
        cell = ws.cell(row=title_row, column=1)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center")

        # Header
        ws.append(df.columns.tolist())
        header_row = ws.max_row
        for col_num, _ in enumerate(df.columns, 1):
            c = ws.cell(row=header_row, column=col_num)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="D9E1F2")
            c.alignment = Alignment(horizontal="center")
            thin = Side(border_style="thin", color="000000")
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Data
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)

        ws.append([])  # Empty row between tables

# Adjust column widths
for col_cells in ws.columns:
    max_length = 0
    col_letter = None
    for cell in col_cells:
        if not isinstance(cell, MergedCell):
            col_letter = cell.column_letter
            break
    if not col_letter:
        continue
    for cell in col_cells:
        if cell.value and not isinstance(cell, MergedCell):
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length + 2

wb.save(buffer)
st.download_button(
    label="Download Excel",
    data=buffer,
    file_name="break_schedule.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
